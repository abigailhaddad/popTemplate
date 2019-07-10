#!/usr/bin/env python
# coding: utf-8

# In[16]:


import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows




###Sample Data
tableIT=pd.DataFrame([ ['DLA', 34, 10000], ['DHRA', 10, 500 ]])
tableServices=pd.DataFrame([ ['DLA', 12, 1000], ['DHRA', 50, 50011 ]])
tablePersonnel=pd.DataFrame([ ['DLA', 'Civilian', 1000], ['DLA', 'Military', 500 ], ['DLA', 'Total', 1500],
                             ['DHRA', 'Civilian', 11000], ['DHRA', 'Military', 500 ], ['DHRA', 'Total', 11500]])

tableOccupation=pd.DataFrame([ ['DLA', 'Logistics', 50], ['DLA', 'Admin', 500 ], ['DLA', 'Acquisitions', 1500],
                             ['DHRA', 'Logistics', 11000], ['DHRA', 'Admin', 500 ], ['DHRA', 'Acquisitions', 11500]])

### Populate Template with Data and Rename

def populateTemplate(listOfTables, nameDAFA):
    #os.chdir(r"H:\_MyComputer\Documents\dafa Emails")
    os.chdir(r"C:\Users\admin\Documents\Python Scripts\template")
    file='template.xlsx'
    wb = load_workbook(filename = file)
    sheet = wb.active 
    #DO Stuff to this
    dictionaryOfTableNamesAndCells={"Name": "B2", 1: [5, 2], 2: [9, 2], 
                                    3: [13, 3], 4: [18,2]}
    listNumber=0
    for table in listOfTables:
        listNumber=listNumber+1
        table=table.loc[table[0]==nameDAFA].drop(table.columns[0], axis=1)
        cellNumbers=dictionaryOfTableNamesAndCells[listNumber]
        writeDFtoWorkbook(sheet, table, cellNumbers[0], cellNumbers[1])
        print(table)
    
    
    sheet[dictionaryOfTableNamesAndCells["Name"]]= nameDAFA
    
    
    
    wb.save(f"{nameDAFA}_template.xlsx")
    


def writeDFtoWorkbook(ws, df, rowNumber, columnNumber):
    rows = dataframe_to_rows(df, index=False, header=False)
    for r_idx, row in enumerate(rows, rowNumber):  
        for c_idx, value in enumerate(row, columnNumber):
             ws.cell(row=r_idx, column=c_idx, value=value)
                
populateTemplate([tableIT, tableServices, tablePersonnel, tableOccupation], "DHRA")


# In[ ]:




