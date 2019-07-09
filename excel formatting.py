# -*- coding: utf-8 -*-
"""
Created on Tue Jul  9 10:39:36 2019

@author: HaddadAE
"""

import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows




###Sample Data
tableIT=pd.DataFrame([ ['DLA', 34, 10000], ['DHRA', 10, 500 ]])
tableServices=pd.DataFrame([ ['DLA', 12, 1000], ['DHRA', 50, 50011 ]])
tablePersonnel=pd.DataFrame([ ['DLA', 'Civilian', 1000], ['DLA', 'Military', 500 ], ['DLA', 'Total', 1500],
                             ['DHRA', 'Civilian', 11000], ['DHRA', 'Military', 500 ], ['DHRA', 'Total', 11500]])

tableOccupation=pd.DataFrame([ ['DLA', 'Logistics', 50], ['DLA', 'Admin', 500 ], ['DLA', 'Acquisitions', 1500],
                             ['DHRA', 'Logistics', 11000], ['DHRA', 'Admin', 500 ], ['DHRA', 'Acquisitions', 11500]])

### Populate Template with Data and Rename

def populateTemplate(listOfTables, nameDAFA):
    os.chdir(r"H:\_MyComputer\Documents\dafa Emails")
    file='template.xlsx'
    wb = load_workbook(filename = file)
    #DO Stuff to this
    dictionaryOfTableNamesAndCells={"Name": "B2", "IT": "B5", "Services": "B9", "Personnel": "C13", "Occupation": "B18"}
    for table in listOfTables:
        table=table.loc[table[0]==nameDAFA]
        print(table)
    
    sheet = wb.active 
    sheet[dictionaryOfTableNamesAndCells["Name"]]= nameDAFA
    
    
    
    wb.save(f"{nameDAFA}_template.xlsx")
    
populateTemplate([tableIT, tableServices, tablePersonnel, tableOccupation], "DHRA")

#def writeDFtoWorkbook()